import openpyxl
import datetime
from CalendarAPI import createEvent, createCalendar, deleteCalendar
import warnings
import json
import time
from datetime import time


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

path = '../Uurroosters/Uurrooster januari 2026.xlsm'

wb_obj = openpyxl.load_workbook(path, data_only=True)

def convert_to_time(value):
    """ Convert different types of values (timedelta, time, or string) to time """
    if isinstance(value, datetime.timedelta):
        # Convert timedelta to time
        total_seconds = value.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        return datetime.time(hours, minutes, seconds)
    elif isinstance(value, datetime.time):
        return value  # Already a time object
    elif isinstance(value, float) or isinstance(value, int):
        # If the value is a number, it might represent an Excel time as a fraction of a day
        # Excel represents time as a fraction of a day (1.0 = 24 hours), we can convert it
        total_seconds = value * 24 * 3600  # Convert days to seconds
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        return datetime.time(hours, minutes, seconds)
    elif isinstance(value, str):
        # Try parsing the string into a time format
        try:
            return datetime.datetime.strptime(value, '%H:%M:%S').time()  # Example: '13:00:00'
        except ValueError:
            return None  # If the string is not in time format, return None
    else:
        return None  # If it's neither time nor timedelta nor a recognizable format


def processShiftsOfColleague(name):
    calendar_ids = None
    try:
        with open("../Calendar_IDs.json", "r", encoding="utf-8") as f:
            calendar_ids = json.load(f)
    except FileNotFoundError:
        print("Calendar_IDs.json not found")
        return

    if calendar_ids is None:
        print("Calender Ids file not found")
        return

    # if name in calendar_ids:
    #     return

    if name not in calendar_ids:
        try:
            cal_id = createCalendar(name)
        except Exception as e:
            return
        calendar_ids[name] = cal_id
        print("Created calendar for " + name + " with id " + cal_id)
        with open("../Calendar_IDs.json", "w", encoding="utf-8") as f:
            json.dump(calendar_ids, f, indent=4)

    cal_id = calendar_ids[name]

    print("Reading sheets")
    for sheet in wb_obj.worksheets:
        print(f"Processing sheet: {sheet.title}")
        if not 'Week' in sheet.title:
            continue

        # Search rows for Person
        personRow = None
        if name != "All":
            for row in range(1, sheet.max_row + 1):
                try:
                    row_name_value = sheet.cell(row=row, column=2).value + " " + sheet.cell(row=row, column=1).value
                except TypeError:
                    continue
                if row_name_value == name:
                    print(f"Found person: {name} on row {row}")
                    personRow = row
                    break

            if personRow is None:
                warnings.warn(f"{name} not found in {sheet.title}")
                continue

        # Get columns with associated dates colleague works
        dates = {}
        for col in range(2, sheet.max_column+1):
            if type(sheet.cell(row=4, column=col).value) is datetime.datetime:
                if name == "All" or sheet.cell(row=personRow, column=col).value is not None:
                    dates[col] = sheet.cell(row=4, column=col).value.replace(year=2026)

        # get exact times of events
        for col in dates:
            if name == "All":
                startTime = time(2, 0)
            else:
                startTime = sheet.cell(row=personRow, column=col).value
            startDate = datetime.datetime.combine(dates[col], startTime)

            if name == "All":
                endTime = time(3, 0)
            else:
                endTime = sheet.cell(row=personRow + 1, column=col).value
            endDate = datetime.datetime.combine(dates[col], endTime)

            if endDate < startDate:
                endDate = endDate + datetime.timedelta(days=1)

            print(str(dates[col]) + ': ' + str(startTime) + '-' + str(endTime))

            colleagues_info = ''
            for colleague in range(1, sheet.max_row + 1):
                try:
                    colleague_name = sheet.cell(row=colleague, column=2).value + sheet.cell(row=colleague, column=1).value[0]
                except TypeError:
                    continue

                startTimeColleague = sheet.cell(row=colleague, column=col).value
                endTimeColleague = sheet.cell(row=colleague + 1, column=col).value
                noteColleague = sheet.cell(row=colleague + 5, column=col).value

                if startTimeColleague is None:
                    continue

                colleagues_info += colleague_name + ": " + str(startTimeColleague) + " - " + str(endTimeColleague)
                if noteColleague is not None:
                    colleagues_info += " (" + str(noteColleague) + ")"
                colleagues_info += ",\n"
            # time.sleep(0.1)
            createEvent(startDate, endDate, cal_id, colleagues_info)

def retrieveColleagues():
    all_colleagues = []
    name_sheet = wb_obj["Maand-gepland"]
    for row in range(4, name_sheet.max_row + 1):
        if name_sheet.cell(row=row, column=1).value in [None, 0, ""]:
            break
        name = name_sheet.cell(row=row, column=2).value + " " + name_sheet.cell(row=row, column=1).value
        all_colleagues.append(name)
    return all_colleagues

def removeCalendars():
    with open("../Calendar_IDs.json", "r", encoding="utf-8") as f:
        calendar_ids = json.load(f)

    for name in list(calendar_ids.keys()):
        deleteCalendar(calendar_ids[name])
        calendar_ids.pop(name)

    with open("../Calendar_IDs.json", "w", encoding="utf-8") as f:
        json.dump(calendar_ids, f, indent=4)

if __name__ == '__main__':
    # removeCalendars()
    # processShiftsOfColleague("Niels Van den Broeck")

    colleagues = retrieveColleagues()
    for colleague in colleagues:
        processShiftsOfColleague(colleague)

    processShiftsOfColleague("All")
    #getAllWorkingPeople()