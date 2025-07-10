import openpyxl
import datetime
from create_event import createEvent
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

first_name = 'All'
last_name = ''
path = 'Uurroosters/Uurrooster julii 2025.xlsm'

wb_obj = openpyxl.load_workbook(path, data_only=True)

def convert_to_time(value):
    """ Convert different types of values (timedelta, time, or string) to time """
    if isinstance(value, datetime.timedelta):
        # Convert timedelta to time
        total_seconds = value.total_seconds()
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        return datetime.time(hours, minutes, seconds)
    elif isinstance(value, datetime.time):
        return value  # Already a time object
    elif isinstance(value, float) or isinstance(value, int):
        # If the value is a number, it might represent an Excel time as a fraction of a day
        # Excel represents time as a fraction of a day (1.0 = 24 hours), we can convert it
        total_seconds = value * 24 * 3600  # Convert days to seconds
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        return datetime.time(hours, minutes, seconds)
    elif isinstance(value, str):
        # Try parsing the string into a time format
        try:
            return datetime.datetime.strptime(value, '%H:%M:%S').time()  # Example: '13:00:00'
        except ValueError:
            return None  # If the string is not in time format, return None
    else:
        return None  # If it's neither time nor timedelta nor a recognizable format



def getDatesOfPerson():
    print("Reading sheets")
    for sheet in wb_obj.worksheets:
        print(f"Processing sheet: {sheet.title}")
        if not 'Week' in sheet.title:
            continue
        # Search rows for Person
        personRow = None
        for row in range(1, sheet.max_row + 1):
            string1 = sheet.cell(row=row, column=1).value  # First name column
            string2 = sheet.cell(row=row, column=2).value  # Last name column

            if string1 is None or string2 is None or isinstance(string1, datetime.time) or isinstance(string2, datetime.time):
                continue

            if first_name in str(string2) and last_name in str(string1):
                print(f"{first_name} {last_name} found in sheet {sheet.title} at row {row}")
                personRow = row
                break

        if personRow is None:
            warnings.warn(f"{first_name} not found in {sheet.title}")
            continue

        # Get columns with associated dates
        dates = {}
        for col in range(2, sheet.max_column+1):
            if(type(sheet.cell(row=4, column=col).value) is datetime.datetime):
                dates[col] = sheet.cell(row=4, column=col).value.replace(year=2025)

        # match columns with row of person and get exact times of events
        for col in dates:
            startTime = sheet.cell(row=personRow, column=col).value
            endTime = sheet.cell(row=personRow + 1, column=col).value
            if startTime is None:
                continue
            print(str(dates[col]) + ': ' + str(startTime) + '-' + str(endTime))
            startDate = datetime.datetime.combine(dates[col], startTime)
            endDate = datetime.datetime.combine(dates[col], endTime)
            colleagues = ''
            for colleague in range(1, sheet.max_row + 1):
                name = sheet.cell(row=colleague, column=2).value
                if name is None or sheet.cell(row=colleague, column=1).value is None or isinstance(name, datetime.time):
                    continue
                startTimeColleague = sheet.cell(row=colleague, column=col).value
                endTimeColleague = sheet.cell(row=colleague + 1, column=col).value
                if startTimeColleague is None:
                    continue
                if str(endTimeColleague) in ['13:00:00', '17:30:00', '18:00:00']  and str(startTime) == '18:30:00':
                    continue
                if str(endTime) in ['13:00:00', '17:30:00', '18:00:00'] and str(startTimeColleague) == '18:30:00':
                    continue
                if sheet.cell(row=colleague + 5, column=col).value == "VIP":
                    name += ' (vip)'
                colleagues += name + ', '

            createEvent(startDate, endDate, first_name, colleagues)


def getAllWorkingPeople():
    for sheet in wb_obj.worksheets:
        for col in range(1, sheet.max_column+1):
            if(type(sheet.cell(row=4, column=col).value) is datetime.datetime):
                date = sheet.cell(row=4, column=col).value.replace(year=2025)
                startTime = datetime.datetime.strptime('13:00:00', '%H:%M:%S').time()
                endTime = datetime.datetime.strptime('23:00:00', '%H:%M:%S').time()
                startDate = datetime.datetime.combine(date, startTime)
                endDate = datetime.datetime.combine(date, endTime)
                colleagues = ''
                for colleague in range(1, sheet.max_row + 1):
                    name = str(sheet.cell(row=colleague, column=2).value)
                    if name is None or sheet.cell(row=colleague, column=1).value is None or isinstance(name, datetime.time):
                        continue
                    startTimeColleague = sheet.cell(row=colleague, column=col).value
                    endTimeColleague = sheet.cell(row=colleague + 1, column=col).value
                    if startTimeColleague is None:
                        continue
                    name += '(' + str(startTimeColleague) + ' tot ' + str(endTimeColleague) + ') '
                    if sheet.cell(row=colleague + 5, column=col).value == "vip":
                        name += ' (vip)'
                    colleagues += name + ', '
                if colleagues == '':
                    continue
                print(str(date))
                createEvent(startDate, endDate, first_name, colleagues)


if __name__ == '__main__':
    if first_name == "All":
        getAllWorkingPeople()
    else:
        getDatesOfPerson()